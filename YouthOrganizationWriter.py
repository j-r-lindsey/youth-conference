import openpyxl
import random

class YouthOrganizationWriter(object):

	SPANISH = "SPANISH"
	ENGLISH = "ENGLISH"

	NORTH_AMERICA = "NORTH AMERICA"
	SOUTH_AMERICA = "SOUTH AMERICA"
	EUROPE = "EUROPE"
	ASIA_PACIFIC = "ASIA PACIFIC (OCEANA)"

	missions = [
		["MEXICO MEXICO CITY","SPANISH",NORTH_AMERICA],
		["ENGLAND LONDON",ENGLISH,EUROPE],
		["FRANCE PARIS","FRENCH",EUROPE],

		["PERU LIMA",SPANISH,SOUTH_AMERICA],
		["CALIFORNIA LOS ANGELES",ENGLISH,NORTH_AMERICA],
		["BRAZIL SAO PAULO","PORTUGUESE", SOUTH_AMERICA],
		
		["ARGENTINA BUENOS AIRES",SPANISH,SOUTH_AMERICA],
		["CANADA TORONTO",ENGLISH,NORTH_AMERICA],
		["PHILLIPPINES MANILLA","TAGALOG",ASIA_PACIFIC],

		["URUGUAY MONTEVIDEO",SPANISH,SOUTH_AMERICA],
		["ILLINOIS CHICAGO",ENGLISH,NORTH_AMERICA],
		["GERMANY BERLIN","GERMAN",EUROPE],
		
		["SPAIN BARCELONA",SPANISH,EUROPE],
		["AUSTRALIA SYDNEY",ENGLISH,ASIA_PACIFIC],
		["ITALY ROME","ITALIAN",EUROPE],
		
		["ECUADOR QUITO",SPANISH,SOUTH_AMERICA],
		["SINGAPORE",ENGLISH,ASIA_PACIFIC],
		["JAPAN TOKYO","JAPANESE",ASIA_PACIFIC]
	]

	def __init__(self,chaperones,hosts):
		self.originalChaperones = chaperones
		self.chaperones = None
		self.originalHosts = chaperones
		self.hosts = None
		self.mission_index = 0

	def getMission(self):
		mission = self.missions[ self.mission_index % len(self.missions) ]
		self.mission_index = self.mission_index + 1
		return mission


	def getChaperone(self, district):
		classQuorum = district[0][3]
		gender = 'Female'
		if classQuorum in ['Priest', 'Teacher']:
			gender = 'Male'

		potentials = [elem for elem in self.chaperones if elem[2] == gender]
		if len(potentials) > 0:
			chap = random.choice(potentials)
			if chap is not None:
				self.chaperones.remove(chap)
				return chap

		return None

	def getHost(self, district):
		# TODO add gender matching
		if len(self.hosts) > 0 :
			return self.hosts.pop()
		else:
			return None


	def writeHeader(self,sheet):
		sheet.cell(row=1, column=1).value = "Group"
		sheet.cell(row=1, column=2).value = "Zone"
		sheet.cell(row=1, column=3).value = "District"
		sheet.cell(row=1, column=4).value = "Mission"
		sheet.cell(row=1, column=5).value = "Language"
		sheet.cell(row=1, column=6).value = "Region"
		sheet.cell(row=1, column=7).value = "Quorum or Class"
		sheet.cell(row=1, column=8).value = "Youth 1"
		sheet.cell(row=1, column=9).value = "Youth 2"
		sheet.cell(row=1, column=10).value = "Youth 3"
		sheet.cell(row=1, column=11).value = "Preference"
		sheet.cell(row=1, column=12).value = "Chaperone"
		sheet.cell(row=1, column=13).value = "Host"
		sheet.cell(row=1, column=14).value = "Notes"

	def writeLargeGroup(self,group,name, sheet):
		i = 1
		for zone in group:
			self.writeZone(zone, name, i, sheet)
			i = i + 1

	def writeZone(self, zone, groupName, zoneNumInGroup, sheet):
		i = 1
		for district in zone:
			self.writeDistrict(district, i, groupName, zoneNumInGroup, sheet)
			i = i + 1

	def writeDistrict(self, district, distNumInZone, groupName, zoneNumInGroup, sheet):
		chaperone = self.getChaperone(district)
		host = self.getHost(district)
		mission = self.getMission()
		for comp in district:
			self.writeCompanionship(comp,distNumInZone, zoneNumInGroup, groupName, mission, chaperone, host, sheet)

	def writeCompanionship(self, comp, distNumInZone, zoneNumInGroup, groupName, mission, chaperone, host, sheet):
		row = sheet.max_row + 1
		sheet.cell(row=row, column=1).value = groupName
		sheet.cell(row=row, column=2).value = zoneNumInGroup
		sheet.cell(row=row, column=3).value = distNumInZone
		sheet.cell(row=row, column=4).value = mission[0] + ' ' + groupName
		sheet.cell(row=row, column=5).value = mission[1]
		sheet.cell(row=row, column=6).value = mission[2]
		sheet.cell(row=row, column=7).value = comp[3]
		sheet.cell(row=row, column=8).value = comp[0]
		sheet.cell(row=row, column=9).value = comp[1]
		sheet.cell(row=row, column=10).value = ""
		sheet.cell(row=row, column=11).value = comp[2]
		sheet.cell(row=row, column=12).value = self.formatAdult(chaperone)
		sheet.cell(row=row, column=13).value = self.formatAdult(host)
		sheet.cell(row=row, column=14).value = ""

	def formatAdult(self,adult):
		if adult is not None:
			return adult[0] + ' ' + adult[1]
		else:
			return ""


	def writeOrganizationToSpreadsheet(self,largeGroups,xlsxPath):
		self.mission_index = 0
		self.chaperones = list(self.originalChaperones)
		self.hosts = list(self.originalHosts)
		wb = openpyxl.Workbook()
		sheet = wb.worksheets[0]
		self.writeHeader(sheet)
		self.writeLargeGroup(largeGroups[0], 'NORTH', sheet)
		self.writeLargeGroup(largeGroups[1], 'SOUTH', sheet)
		self.writeLargeGroup(largeGroups[2], 'WEST', sheet)
		wb.save(xlsxPath)