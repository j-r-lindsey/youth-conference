import openpyxl
import random
import difflib
from VolunteerLoader import VolunteerLoader
from YouthOrganizationWriter import YouthOrganizationWriter

adult_loader = VolunteerLoader('C:/Users/lindsey/Documents/youth-conf/Adult Registration.xlsx')
wb = openpyxl.load_workbook('C:/Users/lindsey/Documents/youth-conf/Youth Registration.xlsx')

# columns
FIRST_NAME = 2
LAST_NAME = 3
GENDER = 5
PREFERENCE_1 = 13
PREFERENCE_2 = 14
PREFERENCE_3 = 19
CLASS_QUORUM = 18

registry = []
assignments = []
available = []
names = []

def getAvailableCompanions():
	comps = []
	for youth in registry:
		if not hasCompanion(youth):
			comps.append(youth)

	return comps

def findYouthByName(fullName): 
	for youth in registry:
		youthName = (youth[0] + ' ' + youth[1]).upper()
		if fullName.strip() == youthName:
			return youth

	return None

def hasCompanion(youth):
	return len(youth) > 7

def createCompanionship(comp1,comp2,comment):
	if not hasCompanion(comp1) and not hasCompanion(comp2):
		comp1.append(comp2[0] + ' ' + comp2[1])
		comp2.append(comp1[0] + ' ' + comp1[1])
		assignments.append([comp2[7],comp1[7],comment,comp1[6]])
		# print("\nCOMPANIONSHIP: " + comp2[7] + ', ' + comp1[7] + '\n')
	else:
		print("ERROR: Already assigned")

def isPreferedByCompToo(youth,comp,depth):
	i = 1
	youthName = (youth[0] + ' ' + youth[1]).upper()
	for pref in [comp[2],comp[3],comp[4]]:
		if i > depth:
			break
			
		if pref is not None:
			if pref.upper().strip() == youthName.strip():
				return True

			alts = difflib.get_close_matches(pref.upper(),names)
			if alts is not None and len(alts) > 0 and alts[0].strip() == youthName.strip():
				return True

		i = i + 1

	return False


def getDoublePreferenceCompanion(youth,depth):
	companion = None
	i = 1
	for pref in [youth[2],youth[3],youth[4]]:
		if i > depth:
			break

		if pref is not None:
			note = None
			comp = findYouthByName(pref.upper())
			if comp is None:
				alts = difflib.get_close_matches(pref.upper(),names)
				if alts is not None and len(alts) > 0:
					# print(pref + " => " + alts[0])
					comp = findYouthByName(alts[0])

			if comp is not None:
				if not hasCompanion(comp):
					if youth[6] == comp[6]:
						if isPreferedByCompToo(youth,comp,depth):
							companion = comp
							break;
						else:
							note = "not in companion's list of preferences."
					else:
						note = "not in same quorum or class."
				else:
					note = "companion already assigned."
			else:
				note = "cannot find in registry"
				# print("Missing youth: " + pref)

			# print('PASS 1: ' + youth[0] + ' ' + youth[1] + ' and ' + pref + ': ' + note)
		i = i + 1

	return companion

def getSinglePreferenceCompanion(youth):
	companion = None
	for pref in [youth[2],youth[3],youth[4]]:
		if pref is not None:
			note = None
			comp = findYouthByName(pref.upper())
			if comp is not None:
				if not hasCompanion(comp):
					if youth[6] == comp[6]:
						companion = comp
						break
					else:
						note = "not in same quorum or class."
				else:
					note = "companion already assigned."
			else:
				note = "cannot find in registry"
				# print("Missing youth: " + pref)

			# print('PASS 2: ' + youth[0] + ' ' + youth[1] + ' and ' + pref + ': ' + note)

	return companion	

def getRandomCompanion(youth):
	comp = random.choice([elem for elem in available if elem[6] == youth[6]])
	if comp is None or comp is youth or not comp[6] == youth[6] or not comp[5] == youth[5]:
		return None
	else:
		return comp

def getDistricts(comps):
	compList = list(comps)
	districts = []
	while len(compList) > 0:
		comps_in_district = []
		for i in range(0, 3):
			options = [elem for elem in compList if not elem in comps_in_district]
			if len(options) > 0:
				comp = random.choice(options)
				compList.remove(comp)
				comps_in_district.append(comp)

		if len(comps_in_district) == 1:
			comp = districts[len(districts)-1].pop()
			comps_in_district.append(comp)

		districts.append(comps_in_district)

	return districts

def countYouthInZone(zone):

	count = 0
	for dist in zone:
		count = count + (len(dist)*2)

	return count

def distributeZone(zone,zoneList):
	while len(zone) > 0:
		smallestZone = None
		smallestSize = 1000
		for z in zoneList:
			size = countYouthInZone(z)
			if size < smallestSize:
				smallestSize = size;
				smallestZone = z

		smallestZone.append(zone.pop())

def getZones(priests,laurels,teachers,miamaids):
	p = list(priests)
	l = list(laurels)
	t = list(teachers)
	m = list(miamaids)
	zones = []
	while True:
		zone = []
		for d in [p,l,t,m]:
			if len(d) > 0:
				comp = random.choice(d)
				d.remove(comp)
				zone.append(comp)
		zlen = len(zone)
		if zlen == 0:
			break
		elif zlen > 2:
			zones.append(zone)
		else:
			distributeZone(zone,zones)

	return zones

def printZone(z):
	print("\nZONE (%d)" % countYouthInZone(z))
	for d in z:
		print('\n')
		for comps in d:
			print(comps)

def printLargeGroup(group):
	print("\nLARGE GROUP =============================================================")
	groupCount = 0
	for z in group:
		zoneCount = countYouthInZone(z)
		groupCount = groupCount + zoneCount
		print("\nZONE (%d)" % zoneCount)
		for d in z:
			print('\n')
			for comps in d:
				print(comps)

	print("Group Total: %d youth in %d zones." % (groupCount,len(group)))

def countReferencesInOrganization(youth,groups):
	count = 0
	name = youth[0] + ' ' + youth[1]
	for group in groups:
		for zone in group:
			for district in zone:
				for comp in district:
					for person in comp:
						if person == name:
							count = count + 1
	return count

def validate(largeGroups):
	for youth in registry:
		count = countReferencesInOrganization(youth,largeGroups)
		if not count == 1:
			print("Validation Error: %s %s found %d times. (%s)" % (youth[0],youth[1],count,str(youth)))

sheet = wb.get_sheet_by_name('Form Responses 1')

for i in range(5, sheet.max_row+1):
	youth = []
	youth.append(sheet.cell(row=i, column=FIRST_NAME).value.strip().upper())
	youth.append(sheet.cell(row=i, column=LAST_NAME).value.strip().upper())
	youth.append(sheet.cell(row=i, column=PREFERENCE_1).value)
	youth.append(sheet.cell(row=i, column=PREFERENCE_2).value)
	youth.append(sheet.cell(row=i, column=PREFERENCE_3).value)
	youth.append(sheet.cell(row=i, column=GENDER).value)
	youth.append(sheet.cell(row=i, column=CLASS_QUORUM).value)
	# if youth[6] == classOrQ:

	name = youth[0] + ' ' + youth[1]
	if not name in names:
		registry.append(youth)
		names.append(name)
	else:
		print('Skipping duplicate: ' + name)

# do we want to order by reg or random?
random.shuffle(registry)

# find bad gender entries
for youth in registry:
	classQuorum = youth[6]
	if youth[5] == 'Male':
		if classQuorum == 'Laurel' or classQuorum == 'Miamaid':
			print("Invalid Entry: " + youth[0] + ' ' + youth[1] + ', ' + youth[6] + ', ' + youth[5])
	elif youth[5] == 'Female':
		if classQuorum == 'Priest' or classQuorum == 'Teacher':
			print("Invalid Entry: " + youth[0] + ' ' + youth[1] + ', ' + youth[6] + ', ' + youth[5])

# first pass - both wanted each other
for i in [1,2,3]:
	for youth in registry:
		if not hasCompanion(youth):
			comp = getDoublePreferenceCompanion(youth,i)
			if comp is not None:
				createCompanionship(youth,comp,'Double %d' % i);

# second pass - one wanted the other
for youth in registry:
	if not hasCompanion(youth):
		comp = getSinglePreferenceCompanion(youth)
		if comp is not None:
			createCompanionship(youth,comp,'Single');

# third pass random
available = getAvailableCompanions()
for youth in registry:
	if not hasCompanion(youth):
		comp = getRandomCompanion(youth)
		if comp is not None:
			createCompanionship(youth,comp,'Random');
			available = getAvailableCompanions()

# X pass, in order
available = getAvailableCompanions()
while len(available) > 1:
	youth = available.pop()
	for comp in available:
		if comp[5] == youth[5] and comp[6] == youth[6]:
			createCompanionship(youth,comp,'Order')
			available.remove(comp)
			break

# for companionship in assignments:
# 	print(companionship)

for youth in available:
	print("Available: " + youth[0] + ' ' + youth[1] + ', ' + youth[6] + ', ' + youth[5])

priestComps = [elem for elem in assignments if elem[3] == 'Priest']
laurelComps =  [elem for elem in assignments if elem[3] == 'Laurel']
teacherComps = [elem for elem in assignments if elem[3] == 'Teacher']
miamaidComps = [elem for elem in assignments if elem[3] == 'Miamaid']

print("\nTotal: " + str(len(registry)))
print("Assigned: " + str(len(assignments)*2))
print("Available: " + str(len(available)))
print("Companionships: " + str(len(assignments)))
print("\tPriests: %d" % len(priestComps))
print("\tLaurels: %d" % len(laurelComps))
print("\tTeachers: %d" % len(teacherComps))
print("\tMiamaids: %d\n" % len(miamaidComps))

# build districts
priestDistricts = getDistricts(priestComps)
laurelDistricts = getDistricts(laurelComps)
teacherDistricts = getDistricts(teacherComps)
miamaidDistricts = getDistricts(miamaidComps)

#build zones
zones = getZones(priestDistricts,laurelDistricts,teacherDistricts,miamaidDistricts)

north = []
south = []
west = []
largeGroups = [north,south,west]
i = 0
while len(zones) > 0:
	if i == 3:
		i = 0

	largeGroups[i].append(zones.pop())
	i = i + 1

# for grp in largeGroups:
# 	printLargeGroup(grp)

validate(largeGroups)

writer = YouthOrganizationWriter(adult_loader.getChaperones(),adult_loader.getHosts())
writer.writeOrganizationToSpreadsheet(largeGroups,"companionships_output.xlsx")










