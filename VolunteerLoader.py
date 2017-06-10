import openpyxl

class VolunteerLoader(object):

	FIRST_NAME = 3
	LAST_NAME = 4
	GENDER = 5
	JOB =10

	chaperones = []
	hosts = []

	def __init__(self, xlsxPath):
		wb = openpyxl.load_workbook('C:/Users/lindsey/Documents/youth-conf/Adult Registration.xlsx')
		sheet = wb.get_sheet_by_name('Form Responses 1')
		for i in range(1, sheet.max_row):
			chap = []
			chap.append(sheet.cell(row=i, column=self.FIRST_NAME).value.strip().upper())
			chap.append(sheet.cell(row=i, column=self.LAST_NAME).value.strip().upper())
			chap.append(sheet.cell(row=i, column=self.GENDER).value)

			if 'Chaperone' in sheet.cell(row=i, column=self.JOB).value:
				self.chaperones.append(chap)

			if 'Host' in sheet.cell(row=i, column=self.JOB).value:
				self.hosts.append(chap)

		wb.close()

	def getChaperones(self):
		return self.chaperones

	def getHosts(self):
		return self.hosts
